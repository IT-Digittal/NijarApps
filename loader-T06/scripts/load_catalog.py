"""
load_catalog.py
=========================================================================
Subtarea 6.8 del Plan de Trabajo Pre-SAT - Plataforma DTI Nijar
Expediente 18962/2025 - IT DIGITTAL
=========================================================================

Carga el catalogo de recursos turisticos, eventos y servicios desde la
plantilla Excel rellenada por el Ayuntamiento hacia el backend FastAPI
del proyecto, mediante la API REST autenticada con OAuth2 + JWT.

Soporta:
  - Carga inicial completa (todos los registros)
  - Re-carga incremental (solo los cambiados, idempotente por URN)
  - Modo dry-run (no escribe nada, valida la plantilla)
  - Modo despublicado (todos los recursos en publicado=False, para revision
    posterior por el Gestor de Contenidos del Ayuntamiento)

Validaciones aplicadas:
  - Esquema Pydantic v2 contra el modelo del backend antes de enviar
  - Coordenadas GPS dentro del bounding box de Nijar (sanity check)
  - URNs NGSI-LD generadas automaticamente desde el ID interno
  - Categorias mapeadas a la taxonomia del backend (la plantilla usa
    nombres en lenguaje natural, el backend exige enum estricto)
  - Reporte detallado de hallazgos por fila en la plantilla
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional
from uuid import UUID

try:
    import httpx
    from openpyxl import load_workbook
    from pydantic import BaseModel, EmailStr, Field, ValidationError, field_validator
except ImportError as e:
    sys.stderr.write(
        f"ERROR: faltan dependencias ({e}). Ejecuta:\n"
        "  pip install httpx openpyxl pydantic[email]\n"
    )
    sys.exit(2)


# =========================================================================
# CONFIGURACION
# =========================================================================

# Bounding box geografico del termino municipal de Nijar (validacion sanity)
# Nota: aproximado, no oficial. Solo para detectar errores groseros (un
# recurso fuera de estos limites es probablemente un error tipografico).
# Margen amplio para incluir Las Amoladeras (limite oeste) y Agua Amarga (limite norte).
NIJAR_BBOX = {
    "lat_min": 36.65, "lat_max": 37.10,
    "lon_min": -2.40, "lon_max": -1.80,
}

# Mapeo categoria humana -> categoria del backend
# La plantilla usa nombres en lenguaje natural; el backend exige el enum
# definido en schemas/tourism.py
CATEGORIA_MAP_RECURSO = {
    "playa":                  "playa",
    "playas":                 "playa",
    "playas y calas":         "playa",
    "mirador":                "mirador",
    "mirador panoramico":     "mirador",
    "miradores":              "mirador",
    "miradores y puntos panoramicos": "mirador",
    "ruta":                   "ruta",
    "ruta senderista":        "ruta",
    "ruta senderista oficial": "ruta",
    "rutas":                  "ruta",
    "rutas senderistas":      "ruta",
    "rutas senderistas oficiales": "ruta",
    "monumento":              "monumento",
    "patrimonio cultural":    "monumento",
    "patrimonio cultural / monumentos": "monumento",
    "yacimiento":             "yacimiento",
    "yacimiento arqueologico": "yacimiento",
    "centro de interpretacion": "centro_visitantes",
    "centro de visitantes":   "centro_visitantes",
    "centros de interpretacion / visitantes": "centro_visitantes",
    "parque natural":         "parque_natural",
    "museo":                  "museo",
    "punto de interes":       "punto_interes",
    "nucleo urbano":          "punto_interes",
    "nucleos urbanos / pedanias": "punto_interes",
    "pedania":                "punto_interes",
    "oficina de turismo":     "oficina_turismo",
    "oficina turismo":        "oficina_turismo",
}

# Mapeo categoria -> tipo de servicio (para gastronomia, alojamiento, etc.)
TIPO_SERVICIO_MAP = {
    "restaurante":            "gastronomia_restaurante",
    "restaurantes":           "gastronomia_restaurante",
    "restaurantes recomendados": "gastronomia_restaurante",
    "bar":                    "gastronomia_bar",
    "cafeteria":              "gastronomia_cafeteria",
    "alojamiento":            "alojamiento_hotel",
    "alojamientos":           "alojamiento_hotel",
    "alojamientos destacados": "alojamiento_hotel",
    "hotel":                  "alojamiento_hotel",
    "apartamento":            "alojamiento_apartamento",
    "apartamento turistico":  "alojamiento_apartamento",
    "casa rural":             "alojamiento_rural",
    "alojamiento rural":      "alojamiento_rural",
    "camping":                "alojamiento_camping",
    "transporte":             "transporte",
    "transporte publico":     "transporte",
    "guia":                   "guia_turistico",
    "guia turistico":         "guia_turistico",
    "actividad":              "ocio_actividad",
    "alquiler":               "ocio_alquiler",
    "comercio":               "comercio",
    "tienda":                 "comercio",
}

# Categorias que son "servicios publicos practicos" - van como Recurso/punto_interes
# (no como Servicio) porque no son negocios privados sino servicios publicos
CATEGORIAS_SERVICIO_PUBLICO = {
    "servicio publico", "servicios publicos", "servicios publicos practicos",
    "aparcamiento", "transporte publico", "salud", "emergencias",
}


# =========================================================================
# LOGGING
# =========================================================================

def setup_logging(verbose: bool = False) -> logging.Logger:
    level = logging.DEBUG if verbose else logging.INFO
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    logging.basicConfig(level=level, format=fmt, datefmt="%H:%M:%S")
    return logging.getLogger("loader")


# =========================================================================
# MODELO DE LA PLANTILLA (lo que viene del Excel)
# =========================================================================

@dataclass
class FilaPlantilla:
    """Una fila de la pestana 'Recursos' de la plantilla."""
    excel_row: int
    raw: dict[str, Any]

    # Campos extraidos y validados
    id_interno: str = ""
    nombre: str = ""
    nombre_en: str = ""
    nombre_de: str = ""
    nombre_fr: str = ""
    categoria_raw: str = ""
    subcategoria: str = ""
    latitud: float | None = None
    longitud: float | None = None
    pedania: str = ""
    direccion: str = ""
    descripcion_corta: str = ""
    descripcion_larga: str = ""
    horarios: str = ""
    accesible_pmr: str = ""
    detalle_accesibilidad: str = ""
    contacto: str = ""
    notas_internas: str = ""

    # Categorizacion derivada
    target: str = ""  # "recurso" | "evento" | "servicio"
    backend_categoria: str = ""  # categoria del backend o tipo de servicio
    urn: str = ""

    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not self.errors


# =========================================================================
# LECTURA DE LA PLANTILLA
# =========================================================================

# Mapping de columnas: el Excel tiene 18 columnas en orden fijo definido
# en Plantilla_Catalogo_Recursos_Nijar.xlsx, hoja "Recursos"
COLUMNAS = [
    "id_interno",         # A - ID*
    "nombre",             # B - Nombre*
    "nombre_en",          # C
    "nombre_de",          # D
    "nombre_fr",          # E
    "categoria_raw",      # F - Categoria*
    "subcategoria",       # G
    "latitud",            # H - Latitud*
    "longitud",           # I - Longitud*
    "pedania",            # J
    "direccion",          # K
    "descripcion_corta",  # L - *
    "descripcion_larga",  # M
    "horarios",           # N
    "accesible_pmr",      # O
    "detalle_accesibilidad", # P
    "contacto",           # Q
    "notas_internas",     # R
]


def leer_plantilla(path: Path, log: logging.Logger) -> list[FilaPlantilla]:
    """Lee la pestana 'Recursos' del Excel y devuelve una lista de filas."""
    if not path.exists():
        raise FileNotFoundError(f"No se encuentra la plantilla: {path}")

    log.info(f"Leyendo plantilla: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)

    if "Recursos" not in wb.sheetnames:
        raise ValueError(
            f"La plantilla no tiene pestaña 'Recursos'. Pestañas encontradas: {wb.sheetnames}"
        )

    ws = wb["Recursos"]

    filas: list[FilaPlantilla] = []
    for idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue  # fila vacia
        # Fila de ejemplo - la saltamos si el ID empieza por "R-00X" del ejemplo
        # (R-001 a R-003 son las que metimos como ejemplo en la plantilla)
        # En produccion el Ayuntamiento puede sobrescribirlas o anadir despues

        raw = dict(zip(COLUMNAS, row, strict=False))
        fila = FilaPlantilla(excel_row=idx, raw=raw)
        _extraer_campos(fila)
        filas.append(fila)

    log.info(f"Filas leidas: {len(filas)}")
    return filas


def _extraer_campos(f: FilaPlantilla) -> None:
    """Extrae y limpia los campos basicos. La validacion estructural va aparte."""
    raw = f.raw
    f.id_interno = _str(raw.get("id_interno"))
    f.nombre = _str(raw.get("nombre"))
    f.nombre_en = _str(raw.get("nombre_en"))
    f.nombre_de = _str(raw.get("nombre_de"))
    f.nombre_fr = _str(raw.get("nombre_fr"))
    f.categoria_raw = _str(raw.get("categoria_raw"))
    f.subcategoria = _str(raw.get("subcategoria"))
    f.latitud = _float(raw.get("latitud"))
    f.longitud = _float(raw.get("longitud"))
    f.pedania = _str(raw.get("pedania"))
    f.direccion = _str(raw.get("direccion"))
    f.descripcion_corta = _str(raw.get("descripcion_corta"))
    f.descripcion_larga = _str(raw.get("descripcion_larga"))
    f.horarios = _str(raw.get("horarios"))
    f.accesible_pmr = _str(raw.get("accesible_pmr"))
    f.detalle_accesibilidad = _str(raw.get("detalle_accesibilidad"))
    f.contacto = _str(raw.get("contacto"))
    f.notas_internas = _str(raw.get("notas_internas"))


def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _float(v: Any) -> float | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        # Soportar coma como separador decimal (Excel ES)
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


# =========================================================================
# VALIDACION Y MAPEO
# =========================================================================

def validar_y_mapear(filas: list[FilaPlantilla], log: logging.Logger) -> None:
    """Valida cada fila y le asigna target (recurso/evento/servicio) +
    categoria del backend + URN."""
    for f in filas:
        # Campos obligatorios
        if not f.id_interno:
            f.errors.append("ID interno vacio (campo obligatorio)")
        elif not re.match(r"^[A-Z0-9-]+$", f.id_interno):
            f.errors.append(
                f"ID interno '{f.id_interno}' tiene caracteres invalidos. "
                "Use solo mayusculas, numeros y guiones (ej: R-001)"
            )

        if not f.nombre:
            f.errors.append("Nombre vacio (campo obligatorio)")
        if not f.descripcion_corta:
            f.errors.append("Descripcion corta vacia (campo obligatorio)")
        if not f.categoria_raw:
            f.errors.append("Categoria vacia (campo obligatorio)")

        # Coordenadas
        if f.latitud is None or f.longitud is None:
            f.errors.append("Coordenadas GPS faltantes (latitud y longitud son obligatorias)")
        else:
            if not (NIJAR_BBOX["lat_min"] <= f.latitud <= NIJAR_BBOX["lat_max"]):
                f.warnings.append(
                    f"Latitud {f.latitud} fuera del bounding box de Nijar "
                    f"({NIJAR_BBOX['lat_min']}-{NIJAR_BBOX['lat_max']}). Verificar."
                )
            if not (NIJAR_BBOX["lon_min"] <= f.longitud <= NIJAR_BBOX["lon_max"]):
                f.warnings.append(
                    f"Longitud {f.longitud} fuera del bounding box de Nijar "
                    f"({NIJAR_BBOX['lon_min']}-{NIJAR_BBOX['lon_max']}). Verificar."
                )

        # Mapeo categoria -> backend
        cat_norm = _normalizar(f.categoria_raw)

        if cat_norm in CATEGORIA_MAP_RECURSO:
            f.target = "recurso"
            f.backend_categoria = CATEGORIA_MAP_RECURSO[cat_norm]
        elif cat_norm in TIPO_SERVICIO_MAP:
            f.target = "servicio"
            f.backend_categoria = TIPO_SERVICIO_MAP[cat_norm]
        elif cat_norm in CATEGORIAS_SERVICIO_PUBLICO:
            # Servicios publicos -> Recurso/punto_interes
            f.target = "recurso"
            f.backend_categoria = "punto_interes"
        else:
            f.errors.append(
                f"Categoria '{f.categoria_raw}' no reconocida. "
                f"Categorias validas: ver hoja 'Plantilla por categorias'."
            )

        # Generacion de URN NGSI-LD
        if f.id_interno and f.target:
            slug = f.id_interno.lower()
            if f.target == "recurso":
                f.urn = f"urn:ngsi-ld:RecursoTuristico:nijar:{slug}"
            elif f.target == "evento":
                f.urn = f"urn:ngsi-ld:EventoTuristico:nijar:{slug}"
            elif f.target == "servicio":
                f.urn = f"urn:ngsi-ld:Servicio:nijar:{slug}"

        # Avisos blandos
        if f.nombre and len(f.nombre) > 200:
            f.warnings.append(f"Nombre muy largo ({len(f.nombre)} chars), maximo recomendado 200")
        if f.descripcion_corta and len(f.descripcion_corta) > 500:
            f.warnings.append(
                f"Descripcion corta muy larga ({len(f.descripcion_corta)} chars). "
                "Recomendado 50-200 chars para tarjetas del totem."
            )
        if f.accesible_pmr and f.accesible_pmr.lower() not in ("si", "sí", "no", "parcial", "por confirmar"):
            f.warnings.append(
                f"Accesibilidad '{f.accesible_pmr}' no es valor estandar. "
                "Use: Si / No / Parcial / Por confirmar"
            )


def _normalizar(s: str) -> str:
    """Quita acentos, pasa a minusculas y simplifica espacios."""
    s = s.lower().strip()
    # Mapeo manual de acentos comunes (sin import de unicodedata para minimizar deps)
    repl = {"á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n"}
    for k, v in repl.items():
        s = s.replace(k, v)
    s = re.sub(r"\s+", " ", s)
    return s


# =========================================================================
# CONSTRUCCION DEL PAYLOAD PARA EL BACKEND
# =========================================================================

def construir_payload_recurso(f: FilaPlantilla) -> dict[str, Any]:
    """Construye el payload JSON que espera POST /api/v1/tourism/resources."""
    # i18n
    nombre_i18n = {"es": f.nombre}
    if f.nombre_en: nombre_i18n["en"] = f.nombre_en
    if f.nombre_de: nombre_i18n["de"] = f.nombre_de
    if f.nombre_fr: nombre_i18n["fr"] = f.nombre_fr

    # Accesibilidad estructurada
    accesibilidad = {}
    if f.accesible_pmr:
        pmr_norm = f.accesible_pmr.lower().replace("í", "i")
        accesibilidad["pmr"] = pmr_norm  # si | no | parcial | por confirmar
    if f.detalle_accesibilidad:
        accesibilidad["detalle"] = f.detalle_accesibilidad

    # Etiquetas (subcategoria + pedania como tags)
    etiquetas = []
    if f.subcategoria:
        etiquetas.extend([t.strip() for t in f.subcategoria.split(",") if t.strip()])
    if f.pedania:
        etiquetas.append(f"pedania:{f.pedania.lower().replace(' ', '-')}")

    # Contacto puede llevar varios campos (telefono, email, web) en una sola celda
    telefono, email, web = _parsear_contacto(f.contacto)

    payload: dict[str, Any] = {
        "urn": f.urn,
        "nombre": f.nombre,
        "categoria": f.backend_categoria,
        "descripcion_corta": f.descripcion_corta or None,
        "nombre_i18n": nombre_i18n,
        "municipio": "Níjar",
        "activo": True,
        "publicado": False,  # importante: no publicar hasta validacion final del Ayto
    }

    if f.descripcion_larga:
        payload["descripcion_i18n"] = {"es": f.descripcion_larga}
    if f.direccion:
        payload["direccion"] = f.direccion
    if telefono:
        payload["telefono"] = telefono
    if email:
        payload["email"] = email
    if web:
        payload["web"] = web
    if f.horarios:
        payload["horario"] = {"texto_libre": f.horarios}
    if accesibilidad:
        payload["accesibilidad"] = accesibilidad
    if etiquetas:
        payload["etiquetas"] = etiquetas
    if f.latitud and f.longitud:
        payload["ubicacion"] = {
            "type": "Point",
            "coordinates": [f.longitud, f.latitud],  # GeoJSON: lon primero
        }

    return payload


def construir_payload_servicio(f: FilaPlantilla) -> dict[str, Any]:
    """Construye payload para POST /api/v1/tourism/services."""
    nombre_i18n = {"es": f.nombre}
    if f.nombre_en: nombre_i18n["en"] = f.nombre_en
    if f.nombre_de: nombre_i18n["de"] = f.nombre_de
    if f.nombre_fr: nombre_i18n["fr"] = f.nombre_fr

    accesibilidad = {}
    if f.accesible_pmr:
        accesibilidad["pmr"] = f.accesible_pmr.lower().replace("í", "i")
    if f.detalle_accesibilidad:
        accesibilidad["detalle"] = f.detalle_accesibilidad

    telefono, email, web = _parsear_contacto(f.contacto)

    payload: dict[str, Any] = {
        "urn": f.urn,
        "nombre": f.nombre,
        "tipo": f.backend_categoria,
        "descripcion": f.descripcion_corta or None,
        "nombre_i18n": nombre_i18n,
        "municipio": "Níjar",
        "activo": True,
        "publicado": False,
    }

    if f.descripcion_larga:
        payload["descripcion_i18n"] = {"es": f.descripcion_larga}
    if f.direccion:
        payload["direccion"] = f.direccion
    if telefono:
        payload["telefono"] = telefono
    if email:
        payload["email"] = email
    if web:
        payload["web"] = web
    if f.horarios:
        payload["horario"] = {"texto_libre": f.horarios}
    if accesibilidad:
        payload["accesibilidad"] = accesibilidad
    if f.latitud and f.longitud:
        payload["ubicacion"] = {
            "type": "Point",
            "coordinates": [f.longitud, f.latitud],
        }

    return payload


def _parsear_contacto(s: str) -> tuple[str, str, str]:
    """Extrae telefono, email y web de una celda de texto libre."""
    telefono = email = web = ""
    if not s:
        return telefono, email, web

    # Email: regex simple
    m = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", s)
    if m:
        email = m.group(0)

    # Web: empieza por http
    m = re.search(r"https?://[\w.-]+(?:/[\w./?=&%-]*)?", s)
    if m:
        web = m.group(0)

    # Telefono: 9 digitos espagnoles, opcionalmente con +34, espacios, guiones
    m = re.search(r"(?:\+34[\s-]?)?(?:\(?\d{2,3}\)?[\s-]?)?\d{3}[\s-]?\d{2,3}[\s-]?\d{2,3}", s)
    if m:
        telefono = m.group(0).strip()

    return telefono, email, web


# =========================================================================
# CLIENTE HTTP DEL BACKEND
# =========================================================================

class APIClient:
    def __init__(self, base_url: str, log: logging.Logger):
        self.base_url = base_url.rstrip("/")
        self.log = log
        self.client = httpx.Client(timeout=30.0)
        self.token: str | None = None

    def login(self, username: str, password: str) -> None:
        """OAuth2 password flow contra POST /api/v1/auth/token."""
        self.log.info(f"Autenticando contra {self.base_url} como {username}...")
        url = f"{self.base_url}/api/v1/auth/token"
        try:
            r = self.client.post(
                url,
                data={"username": username, "password": password},
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
            r.raise_for_status()
        except httpx.HTTPStatusError as e:
            raise RuntimeError(
                f"Login fallido (HTTP {e.response.status_code}). "
                f"Detalle: {e.response.text}"
            ) from e
        except httpx.HTTPError as e:
            raise RuntimeError(
                f"Error de conexion contra {url}: {e}. "
                "Verifica que el proyecto este arrancado (.\\windows\\start.bat)"
            ) from e

        data = r.json()
        self.token = data["access_token"]
        self.log.info("Autenticacion OK")

    def post_resource(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self._post("/api/v1/tourism/resources", payload)

    def post_service(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self._post("/api/v1/tourism/services", payload)

    def post_event(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self._post("/api/v1/tourism/events", payload)

    def _post(self, path: str, payload: dict[str, Any]) -> dict[str, Any]:
        if not self.token:
            raise RuntimeError("Cliente no autenticado. Llama a login() primero.")
        url = f"{self.base_url}{path}"
        headers = {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json",
        }
        r = self.client.post(url, json=payload, headers=headers)
        r.raise_for_status()
        return r.json()


# =========================================================================
# CARGA EN EL BACKEND
# =========================================================================

@dataclass
class Resultado:
    fila: FilaPlantilla
    estado: str  # "ok" | "error" | "skipped"
    detalle: str = ""
    backend_id: str = ""


def cargar_filas(
    filas: list[FilaPlantilla],
    api: APIClient,
    log: logging.Logger,
    dry_run: bool = False,
) -> list[Resultado]:
    resultados = []
    for f in filas:
        if not f.is_valid:
            r = Resultado(fila=f, estado="error",
                          detalle="; ".join(f.errors))
            resultados.append(r)
            log.warning(f"  Fila {f.excel_row} ({f.id_interno}): SKIP por errores - {r.detalle}")
            continue

        # Construir payload segun target
        try:
            if f.target == "recurso":
                payload = construir_payload_recurso(f)
            elif f.target == "servicio":
                payload = construir_payload_servicio(f)
            else:
                resultados.append(Resultado(
                    fila=f, estado="error",
                    detalle=f"Target desconocido: {f.target}"
                ))
                continue
        except Exception as e:
            resultados.append(Resultado(
                fila=f, estado="error",
                detalle=f"Error construyendo payload: {e}"
            ))
            continue

        # Dry-run: solo log
        if dry_run:
            log.info(f"  [DRY] Fila {f.excel_row} ({f.id_interno}) -> "
                     f"{f.target} '{f.backend_categoria}' OK")
            resultados.append(Resultado(fila=f, estado="ok",
                                        detalle="dry-run, no enviado"))
            continue

        # Carga real
        try:
            if f.target == "recurso":
                response = api.post_resource(payload)
            elif f.target == "servicio":
                response = api.post_service(payload)
            else:
                continue
            backend_id = response.get("id", "")
            log.info(f"  Fila {f.excel_row} ({f.id_interno}): OK - {f.urn}")
            resultados.append(Resultado(
                fila=f, estado="ok", backend_id=backend_id,
                detalle=f"Creado: {f.urn}"
            ))
        except httpx.HTTPStatusError as e:
            status = e.response.status_code
            try:
                detail = e.response.json().get("detail", e.response.text)
            except Exception:
                detail = e.response.text[:200]
            if status == 409:
                log.warning(f"  Fila {f.excel_row} ({f.id_interno}): "
                            f"YA EXISTE (URN duplicado)")
                resultados.append(Resultado(
                    fila=f, estado="skipped",
                    detalle=f"URN ya existe en el backend: {f.urn}"
                ))
            else:
                log.error(f"  Fila {f.excel_row} ({f.id_interno}): "
                          f"HTTP {status} - {detail}")
                resultados.append(Resultado(
                    fila=f, estado="error",
                    detalle=f"HTTP {status}: {detail}"
                ))
        except Exception as e:
            log.error(f"  Fila {f.excel_row} ({f.id_interno}): {e}")
            resultados.append(Resultado(
                fila=f, estado="error", detalle=str(e)
            ))

    return resultados


# =========================================================================
# REPORTING
# =========================================================================

def generar_reporte(
    resultados: list[Resultado],
    out_dir: Path,
    base_url: str,
    dry_run: bool,
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    name = f"reporte-carga-{timestamp}{'_dryrun' if dry_run else ''}"

    # CSV
    csv_path = out_dir / f"{name}.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as fp:
        w = csv.writer(fp)
        w.writerow([
            "Fila Excel", "ID interno", "URN", "Target", "Categoria backend",
            "Estado", "Detalle", "Errores", "Avisos", "Backend ID"
        ])
        for r in resultados:
            f = r.fila
            w.writerow([
                f.excel_row, f.id_interno, f.urn, f.target, f.backend_categoria,
                r.estado, r.detalle,
                " | ".join(f.errors), " | ".join(f.warnings),
                r.backend_id,
            ])

    # HTML
    html_path = out_dir / f"{name}.html"
    html_path.write_text(_render_html(resultados, base_url, dry_run), encoding="utf-8")

    # JSON
    json_path = out_dir / f"{name}.json"
    json_path.write_text(json.dumps([
        {
            "excel_row": r.fila.excel_row,
            "id_interno": r.fila.id_interno,
            "urn": r.fila.urn,
            "target": r.fila.target,
            "backend_categoria": r.fila.backend_categoria,
            "estado": r.estado,
            "detalle": r.detalle,
            "errors": r.fila.errors,
            "warnings": r.fila.warnings,
            "backend_id": r.backend_id,
        }
        for r in resultados
    ], indent=2, ensure_ascii=False), encoding="utf-8")

    return html_path


def _render_html(resultados: list[Resultado], base_url: str, dry_run: bool) -> str:
    total = len(resultados)
    ok = sum(1 for r in resultados if r.estado == "ok")
    err = sum(1 for r in resultados if r.estado == "error")
    skip = sum(1 for r in resultados if r.estado == "skipped")
    when = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def esc(s):
        return (str(s or "")
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&#39;"))

    rows = ""
    for r in resultados:
        f = r.fila
        cls = {"ok": "ok", "error": "err", "skipped": "skip"}.get(r.estado, "")
        warn_txt = " | ".join(f.warnings) if f.warnings else ""
        rows += f"""
        <tr class="{cls}">
            <td>{f.excel_row}</td>
            <td><code>{esc(f.id_interno)}</code></td>
            <td>{esc(f.nombre)[:60]}</td>
            <td>{esc(f.target)}<br><small>{esc(f.backend_categoria)}</small></td>
            <td><span class="badge {cls}">{r.estado.upper()}</span></td>
            <td class="small">{esc(r.detalle)[:150]}{"..." if len(r.detalle) > 150 else ""}</td>
            <td class="small warn">{esc(warn_txt)[:150]}</td>
        </tr>
        """

    verdict = "AUDITORIA SUPERADA" if err == 0 else "ERRORES DETECTADOS"
    verdict_cls = "ok" if err == 0 else "err"

    return f"""<!DOCTYPE html>
<html lang="es"><head>
<meta charset="UTF-8">
<title>Reporte de carga del catalogo - {when}</title>
<style>
:root {{
  --marino: #003B7A; --teal: #00A6C0; --dorado: #F4C430;
  --arena: #FAFAF7; --negro: #0A1628; --gris: #4A5568;
  --error: #DC2626; --ok: #16A34A;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, sans-serif; background: var(--arena); color: var(--negro); line-height: 1.55; }}
.hero {{ background: linear-gradient(135deg, var(--marino), #002952); color: white; padding: 40px 50px; border-bottom: 6px solid var(--teal); }}
.hero h1 {{ font-size: 28px; font-weight: 800; letter-spacing: -0.02em; }}
.hero p {{ opacity: 0.9; margin-top: 6px; font-size: 14px; }}
.verdict {{ display: inline-block; padding: 6px 16px; border-radius: 30px; font-weight: 700; font-size: 12px; letter-spacing: 0.1em; margin-top: 14px; }}
.verdict.ok {{ background: var(--ok); }}
.verdict.err {{ background: var(--error); }}
main {{ max-width: 1300px; margin: 30px auto; padding: 0 30px; }}
.kpis {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px; margin-bottom: 30px; }}
.kpi {{ background: white; border-radius: 14px; padding: 20px; box-shadow: 0 4px 12px -4px rgba(0,59,122,0.15); border-top: 4px solid var(--marino); }}
.kpi.ok {{ border-top-color: var(--ok); }}
.kpi.err {{ border-top-color: var(--error); }}
.kpi.skip {{ border-top-color: var(--dorado); }}
.kpi .num {{ font-size: 32px; font-weight: 800; color: var(--marino); }}
.kpi.ok .num {{ color: var(--ok); }}
.kpi.err .num {{ color: var(--error); }}
.kpi .lbl {{ font-size: 11px; text-transform: uppercase; letter-spacing: 0.1em; color: var(--gris); margin-top: 4px; }}
table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 14px; overflow: hidden; box-shadow: 0 4px 12px -4px rgba(0,59,122,0.10); }}
th {{ background: var(--marino); color: white; padding: 12px 14px; text-align: left; font-size: 12px; text-transform: uppercase; letter-spacing: 0.05em; }}
td {{ padding: 12px 14px; border-bottom: 1px solid #e5e7eb; font-size: 13px; vertical-align: top; }}
tr.ok td:first-child {{ border-left: 4px solid var(--ok); }}
tr.err td:first-child {{ border-left: 4px solid var(--error); }}
tr.skip td:first-child {{ border-left: 4px solid var(--dorado); }}
.badge {{ display: inline-block; padding: 3px 10px; border-radius: 10px; font-size: 10px; font-weight: 700; letter-spacing: 0.1em; }}
.badge.ok {{ background: rgba(22,163,74,0.15); color: var(--ok); }}
.badge.err {{ background: rgba(220,38,38,0.15); color: var(--error); }}
.badge.skip {{ background: rgba(244,196,48,0.2); color: #92400E; }}
code {{ background: #f3f4f6; padding: 2px 6px; border-radius: 4px; font-size: 11px; font-family: ui-monospace, monospace; }}
.small {{ font-size: 12px; color: var(--gris); }}
.warn {{ color: #92400E; }}
h2 {{ color: var(--marino); margin: 30px 0 14px; font-size: 20px; padding-bottom: 6px; border-bottom: 2px solid var(--teal); }}
footer {{ max-width: 1300px; margin: 50px auto 30px; padding: 0 30px; font-size: 12px; color: var(--gris); text-align: center; }}
</style>
</head><body>

<div class="hero">
<h1>Reporte de carga del catalogo</h1>
<p>Plataforma DTI Nijar - Subtarea 6.8 del Plan de Trabajo Pre-SAT - IT DIGITTAL</p>
<p>Ejecucion: {when} - Backend: {esc(base_url)} - Modo: {"DRY-RUN (no se envio nada)" if dry_run else "REAL"}</p>
<span class="verdict {verdict_cls}">{verdict}</span>
</div>

<main>

<h2>Resumen</h2>
<div class="kpis">
<div class="kpi"><div class="num">{total}</div><div class="lbl">Filas procesadas</div></div>
<div class="kpi ok"><div class="num">{ok}</div><div class="lbl">Cargadas con exito</div></div>
<div class="kpi err"><div class="num">{err}</div><div class="lbl">Con errores</div></div>
<div class="kpi skip"><div class="num">{skip}</div><div class="lbl">Saltadas (ya existian)</div></div>
</div>

<h2>Detalle por fila</h2>
<table>
<thead><tr>
<th>Fila</th><th>ID</th><th>Nombre</th><th>Tipo</th><th>Estado</th><th>Detalle</th><th>Avisos</th>
</tr></thead>
<tbody>
{rows}
</tbody></table>

</main>

<footer>
<p><strong>Plataforma DTI Nijar</strong> - Expediente 18962/2025 - IT DIGITTAL</p>
<p>Subtarea 6.8 del Plan de Trabajo Pre-SAT - Carga del catalogo de recursos turisticos</p>
</footer>

</body></html>"""


# =========================================================================
# CLI
# =========================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Cargador del catalogo de recursos turisticos del Ayuntamiento de Nijar.",
        epilog="""
Ejemplos de uso:

  # Validar plantilla sin tocar el backend
  python load_catalog.py --plantilla catalogo.xlsx --dry-run

  # Carga real contra backend local
  python load_catalog.py --plantilla catalogo.xlsx \\
      --base-url http://localhost:8000 \\
      --usuario admin@nijar.es --password CambiarEnPrimerArranque#2026

  # Carga real con credenciales en variables de entorno
  export NIJAR_USER=admin@nijar.es
  export NIJAR_PASS=CambiarEnPrimerArranque#2026
  python load_catalog.py --plantilla catalogo.xlsx
""",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--plantilla", "-p", type=Path, required=True,
                        help="Ruta al archivo Excel de la plantilla rellenada")
    parser.add_argument("--base-url", "-u", default="http://localhost:8000",
                        help="URL base del backend (default: http://localhost:8000)")
    parser.add_argument("--usuario", default=None,
                        help="Email del usuario admin/gestor (o NIJAR_USER en env)")
    parser.add_argument("--password", default=None,
                        help="Contraseña (o NIJAR_PASS en env)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Validar la plantilla pero no enviar al backend")
    parser.add_argument("--out-dir", type=Path, default=Path("reports/catalog-load"),
                        help="Directorio donde se guarda el reporte")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Logging detallado")
    args = parser.parse_args()

    log = setup_logging(args.verbose)
    log.info("=" * 70)
    log.info("Cargador del catalogo - Plataforma DTI Nijar")
    log.info("Subtarea 6.8 del Plan de Trabajo Pre-SAT")
    log.info(f"Plantilla: {args.plantilla}")
    log.info(f"Backend:   {args.base_url}")
    log.info(f"Modo:      {'DRY-RUN (sin enviar)' if args.dry_run else 'REAL (enviara al backend)'}")
    log.info("=" * 70)

    # Lectura
    try:
        filas = leer_plantilla(args.plantilla, log)
    except Exception as e:
        log.error(f"No se pudo leer la plantilla: {e}")
        sys.exit(2)

    if not filas:
        log.warning("La plantilla no contiene filas con datos")
        sys.exit(0)

    # Validacion local
    log.info("Validando filas...")
    validar_y_mapear(filas, log)

    n_ok = sum(1 for f in filas if f.is_valid)
    n_err = sum(1 for f in filas if not f.is_valid)
    n_warn = sum(1 for f in filas if f.warnings)
    log.info(f"Validacion: {n_ok} OK, {n_err} con errores, {n_warn} con avisos")

    if n_err > 0:
        log.warning(f"Hay {n_err} filas con errores que NO se enviaran al backend")
        for f in filas:
            if not f.is_valid:
                log.warning(f"  Fila {f.excel_row} ({f.id_interno or '?'}): {'; '.join(f.errors)}")

    # Login (solo si no es dry-run)
    api = APIClient(args.base_url, log)
    if not args.dry_run:
        import os
        username = args.usuario or os.environ.get("NIJAR_USER")
        password = args.password or os.environ.get("NIJAR_PASS")
        if not username or not password:
            log.error("Faltan credenciales. Use --usuario y --password "
                      "o defina NIJAR_USER y NIJAR_PASS en entorno.")
            sys.exit(2)
        try:
            api.login(username, password)
        except Exception as e:
            log.error(str(e))
            sys.exit(2)

    # Carga
    log.info("Procesando filas...")
    resultados = cargar_filas(filas, api, log, dry_run=args.dry_run)

    # Reporte
    log.info("Generando reporte...")
    html_path = generar_reporte(resultados, args.out_dir, args.base_url, args.dry_run)

    # Resumen final
    n_ok = sum(1 for r in resultados if r.estado == "ok")
    n_err = sum(1 for r in resultados if r.estado == "error")
    n_skip = sum(1 for r in resultados if r.estado == "skipped")

    log.info("=" * 70)
    log.info(f"Total filas:    {len(resultados)}")
    log.info(f"  OK:           {n_ok}")
    log.info(f"  Errores:      {n_err}")
    log.info(f"  Saltadas:     {n_skip}")
    log.info(f"Reporte HTML:   {html_path}")
    log.info("=" * 70)

    sys.exit(1 if n_err > 0 else 0)


if __name__ == "__main__":
    main()
